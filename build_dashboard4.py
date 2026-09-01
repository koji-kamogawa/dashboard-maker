#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SharePoint「サイトの使用状況データ」統合ダッシュボード生成スクリプト
========================================================================

概要
----
毎週ダウンロードした SharePoint の使用状況データ xlsx
（シート構成: 全体的なトラフィック / 人気のあるコンテンツ /
             デバイス別の使用状況 / 時間別の使用状況）
が1つのフォルダにたまっている前提で、それらをすべて読み込み、
- 日付の重複を除去して時系列に統合
- 集計（7日/30日/90日/全期間）
- グラフ（トラフィック推移・デバイス別推移・人気コンテンツ・時間帯別ヒートマップ）
を1枚の自己完結型HTMLダッシュボードとして出力します。

使い方
------
    pip install pandas openpyxl matplotlib
    python build_dashboard.py <ログが入ったフォルダ> [出力HTMLファイル名]

例:
    python build_dashboard.py "C:\\Reports\\glosサイトログ" dashboard.html

前提・注意点
------------
- フォルダ内の .xlsx はすべて同じサイトから毎週エクスポートしたものを想定しています
  （ファイル名は問いません。中身のシート構成だけを見ます）。
- 「全体的なトラフィック」「デバイス別の使用状況」は直近90日分が毎回入っているため、
  複数週のファイルを重ねると日付が重複します。
  → 同じ日付が複数ファイルに存在する場合は、ダウンロード日が最も新しいファイルの値を採用します。
- 「重複しない閲覧者数(UU)」は日をまたいで同一人物が重複しうるため、日次値を単純合計しても
  正しい期間合計にはなりません。このスクリプトでは、各ファイルの「集計データ」表
  （過去7日/過去30日/過去90日間/すべての時間の公式集計値）のうち、最新ファイルの値を
  そのままサマリーとして表示します（日次値の合算はグラフの形状確認にのみ使用）。
- 日本語フォントが環境に無いとグラフの文字が四角(□)になります。その場合は
  Noto Sans CJK JP / IPAexGothic / Meiryo などの日本語フォントをインストールしてください。
"""

import sys
import os
import glob
import re
import json
import math
import datetime
import base64
from io import BytesIO

import pandas as pd
import openpyxl

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.font_manager as fm

# ----------------------------------------------------------------------------
# 日本語フォントの自動選択（無ければ警告を出して既定フォントで続行）
# ----------------------------------------------------------------------------
def setup_japanese_font():
    """日本語フォントを探して matplotlib に設定する。

    手順:
      1) まずフォント名で検索する（ローカルWindows/macOS環境など、
         OSにインストール済みのフォントをそのまま使えるケース向け）。
      2) 名前で見つからない場合、既知のファイルパスを直接
         fm.fontManager.addfont() で読み込む。
         Streamlit Cloud（Debian/Ubuntu）で apt 経由インストールされる
         fonts-noto-cjk は「可変フォント(Variable Font)」形式の .ttc であり、
         matplotlib が使う FreeType がこれを正しく解析できず
         "RuntimeError: broken table" 等で読み込みに失敗する既知の問題があるため、
         静的フォントである IPAフォント／IPAexフォント／Takaoフォントを優先的に
         フォールバック候補として試す。
      3) それでも見つからなければ警告を出し、日本語グリフは文字化け（□）のまま続行する。
    """
    candidates = [
        "Noto Sans CJK JP", "Noto Sans JP", "IPAexGothic", "IPAGothic",
        "Yu Gothic", "Meiryo", "Hiragino Sans", "MS Gothic", "TakaoPGothic",
    ]
    available = {f.name for f in fm.fontManager.ttflist}
    for name in candidates:
        if name in available:
            plt.rcParams["font.family"] = name
            return name

    # --- フォールバック: 既知のパスにある「静的」日本語フォントを直接読み込む ---
    # 可変フォント(Noto Sans CJK の .ttc)は matplotlib で読み込めないことがあるため、
    # ここでは静的フォント（IPA/IPAex/Takao/VLGothic）のパスのみを候補にする。
    fallback_paths = [
        "/usr/share/fonts/opentype/ipaexfont-gothic/ipaexg.ttf",   # fonts-ipaexfont-gothic
        "/usr/share/fonts/opentype/ipafont-gothic/ipag.ttf",       # fonts-ipafont-gothic
        "/usr/share/fonts/truetype/takao-gothic/TakaoPGothic.ttf", # fonts-takao
        "/usr/share/fonts/truetype/fonts-japanese-gothic.ttf",     # 一部ディストリ
        "/usr/share/fonts/truetype/vlgothic/VL-PGothic-Regular.ttf",  # fonts-vlgothic
    ]
    for path in fallback_paths:
        if not os.path.exists(path):
            continue
        try:
            fm.fontManager.addfont(path)
            prop = fm.FontProperties(fname=path)
            font_name = prop.get_name()
            plt.rcParams["font.family"] = font_name
            return font_name
        except Exception as e:
            print(f"[警告] フォント読み込みに失敗しました（{path}）: {e}", file=sys.stderr)
            continue

    print("[警告] 日本語フォントが見つかりませんでした。グラフの日本語が文字化けする場合は "
          "packages.txt に fonts-ipafont-gothic（または fonts-ipaexfont-gothic）を追加してください。",
          file=sys.stderr)
    return None

setup_japanese_font()
plt.rcParams["axes.unicode_minus"] = False


# ----------------------------------------------------------------------------
# ヘルパー: シート内で見出しセルを探し、その下のデータ行を返す
# ----------------------------------------------------------------------------
def find_header_row(ws, header_text, col=1, max_row=200):
    for r in range(1, min(ws.max_row, max_row) + 1):
        v = ws.cell(row=r, column=col).value
        if v == header_text:
            return r
    return None


def parse_download_date(ws, fallback_path):
    """シート冒頭の『レポートのダウンロード日: 27日-7月,2026年』を日付に変換。
    見つからない場合はファイルの更新日時を使う。"""
    for r in range(1, 6):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and "ダウンロード日" in v:
            m = re.search(r"(\d+)日-(\d+)月,(\d+)年", v)
            if m:
                d, mo, y = map(int, m.groups())
                return datetime.date(y, mo, d)
    ts = os.path.getmtime(fallback_path)
    return datetime.date.fromtimestamp(ts)


def parse_download_date_or_default(ws, fallback_date):
    """シート冒頭の『レポートのダウンロード日: 27日-7月,2026年』を日付に変換。
    見つからない場合は fallback_date（呼び出し側から渡された日付）をそのまま使う。
    アップロードされた file-like オブジェクトにはファイル更新日時が無いため、
    parse_download_date() とは別に用意した派生版。"""
    for r in range(1, 6):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and "ダウンロード日" in v:
            m = re.search(r"(\d+)日-(\d+)月,(\d+)年", v)
            if m:
                d, mo, y = map(int, m.groups())
                return datetime.date(y, mo, d)
    return fallback_date


def extract_daily_sheet(ws, value_cols):
    """『日付』ヘッダーを持つシート（全体的なトラフィック / デバイス別の使用状況）から
    日付と数値列を抽出する。value_cols は見出し名のリスト。"""
    hdr_row = find_header_row(ws, "日付", col=1)
    if hdr_row is None:
        return []
    rows = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        d = ws.cell(row=r, column=1).value
        if not isinstance(d, datetime.datetime):
            continue
        vals = [ws.cell(row=r, column=c).value for c in range(2, 2 + len(value_cols))]
        rows.append((d.date(), *vals))
    return rows


def extract_content_sheet(ws):
    hdr_row = find_header_row(ws, "コンテンツ", col=1)
    if hdr_row is None:
        return []
    rows = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        typ = ws.cell(row=r, column=2).value
        uu7 = ws.cell(row=r, column=3).value
        pv7 = ws.cell(row=r, column=4).value
        if name is None or not isinstance(uu7, (int, float)) or not isinstance(pv7, (int, float)):
            continue
        rows.append((name, typ, uu7, pv7))
    return rows


def extract_hourly_sheet(ws):
    hdr_row = find_header_row(ws, "時間", col=1)
    if hdr_row is None:
        return []
    rows = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        avg7 = ws.cell(row=r, column=2).value
        if not isinstance(label, str) or "-" not in label and " " not in label:
            continue
        if not isinstance(avg7, (int, float)):
            continue
        rows.append((label, avg7))
    return rows


def extract_summary_block(ws):
    """『期間』ヘッダーの下にある集計データ(過去7日/過去30日/過去90日間/すべての時間)を取得。"""
    hdr_row = find_header_row(ws, "期間", col=1)
    if hdr_row is None:
        return {}
    out = {}
    for r in range(hdr_row + 1, hdr_row + 6):
        label = ws.cell(row=r, column=1).value
        uu = ws.cell(row=r, column=2).value
        pv = ws.cell(row=r, column=3).value
        if label in ("過去 7 日", "過去 30 日", "過去 90 日間", "すべての時間"):
            out[label] = (uu, pv)
    return out


# ----------------------------------------------------------------------------
# メイン: フォルダ内の全ファイルを読み込む
# ----------------------------------------------------------------------------
def load_all(folder):
    files = sorted(glob.glob(os.path.join(folder, "*.xlsx")))
    if not files:
        raise SystemExit(f"'{folder}' に .xlsx ファイルが見つかりません。")

    traffic_rows, device_rows, content_rows, hourly_rows = [], [], [], []
    summaries = {}
    latest_dl_date, latest_hourly, latest_ws_names = None, None, None

    for path in files:
        wb = openpyxl.load_workbook(path, data_only=True)
        dl_date = None
        if "全体的なトラフィック" in wb.sheetnames:
            ws = wb["全体的なトラフィック"]
            dl_date = parse_download_date(ws, path)
            for d, uu, pv in extract_daily_sheet(ws, ["UU", "PV"]):
                traffic_rows.append((dl_date, d, uu, pv))
            summaries[dl_date] = extract_summary_block(ws)

        if "デバイス別の使用状況" in wb.sheetnames:
            ws = wb["デバイス別の使用状況"]
            if dl_date is None:
                dl_date = parse_download_date(ws, path)
            for row in extract_daily_sheet(ws, ["desktop", "mobileapp", "mobileweb", "tablet", "other"]):
                device_rows.append((dl_date, *row))

        if "人気のあるコンテンツ" in wb.sheetnames:
            ws = wb["人気のあるコンテンツ"]
            if dl_date is None:
                dl_date = parse_download_date(ws, path)
            for name, typ, uu7, pv7 in extract_content_sheet(ws):
                content_rows.append((dl_date, name, typ, uu7, pv7))

        if "時間別の使用状況" in wb.sheetnames:
            ws = wb["時間別の使用状況"]
            if dl_date is None:
                dl_date = parse_download_date(ws, path)
            rows = extract_hourly_sheet(ws)
            if latest_dl_date is None or dl_date >= latest_dl_date:
                latest_hourly = rows

        if latest_dl_date is None or (dl_date and dl_date >= latest_dl_date):
            latest_dl_date = dl_date

        print(f"読み込み: {os.path.basename(path)}  (ダウンロード日: {dl_date})")

    traffic_df = pd.DataFrame(traffic_rows, columns=["取得日", "日付", "UU", "PV"])
    device_df = pd.DataFrame(device_rows, columns=["取得日", "日付", "desktop", "mobileapp", "mobileweb", "tablet", "other"])
    content_df = pd.DataFrame(content_rows, columns=["取得日", "コンテンツ", "種類", "UU7", "PV7"])
    hourly_df = pd.DataFrame(latest_hourly or [], columns=["時間帯", "平均UU_7日"])

    # --- 重複除去: 同じ日付は「取得日」が最も新しい行を残す ---
    def dedup_by_date(df):
        if df.empty:
            return df
        df = df.sort_values(["日付", "取得日"])
        return df.drop_duplicates(subset="日付", keep="last").sort_values("日付").reset_index(drop=True)

    traffic_df = dedup_by_date(traffic_df)
    device_df = dedup_by_date(device_df)

    latest_summary = summaries.get(latest_dl_date, {})

    # --- 欠落期間チェック: 統合後の日付に連続した抜け(欠測日)が無いか ---
    gaps = []
    if not traffic_df.empty:
        dates_sorted = sorted(traffic_df["日付"].unique())
        for prev, cur in zip(dates_sorted, dates_sorted[1:]):
            gap_days = (cur - prev).days
            if gap_days > 1:
                gaps.append((prev, cur, gap_days))
    if gaps:
        print(f"[警告] データに {len(gaps)} 件の抜け(欠測期間)があります。"
              f"90日を超えて次のファイルをダウンロードしなかった期間がある可能性があります。")
        for prev, cur, g in gaps[:10]:
            print(f"   {prev} 〜 {cur} の間 ({g-1}日分欠測)")

    return {
        "traffic": traffic_df,
        "device": device_df,
        "content": content_df,
        "hourly": hourly_df,
        "latest_dl_date": latest_dl_date,
        "latest_summary": latest_summary,
        "n_files": len(files),
        "gaps": gaps,
        "file_warnings": [],
    }


# ----------------------------------------------------------------------------
# Streamlit（アップロード）版: フォルダではなく file-like オブジェクトのlistから読み込む
# ----------------------------------------------------------------------------
def load_all_from_sources(sources):
    """sources: [(ファイル名, file-likeオブジェクト, フォールバック日付), ...] の形式のlist。
    dashboard_maker.py（Streamlit UI）から、st.file_uploader でアップロードされた
    複数の xlsx をこの形式に変換して渡すことを想定している。
    load_all(folder) とほぼ同じロジックだが、
      - openpyxl.load_workbook にファイルパスではなく file-like を渡す
      - 「ダウンロード日」が見つからない場合のフォールバックに os.path.getmtime ではなく
        呼び出し側が渡した日付（today など）を使う
      - 個々のファイルの読み込み失敗やシート欠如を例外で落とさず file_warnings に集約する
    点が異なる。"""
    traffic_rows, device_rows, content_rows, hourly_rows = [], [], [], []
    summaries = {}
    latest_dl_date, latest_hourly = None, None
    file_warnings = []

    for name, file_like, fallback_date in sources:
        try:
            wb = openpyxl.load_workbook(file_like, data_only=True)
        except Exception as e:
            file_warnings.append((name, f"読み込みに失敗しました（xlsx形式を確認してください）: {e}"))
            continue

        dl_date = None
        found_any_sheet = False

        if "全体的なトラフィック" in wb.sheetnames:
            found_any_sheet = True
            ws = wb["全体的なトラフィック"]
            dl_date = parse_download_date_or_default(ws, fallback_date)
            for d, uu, pv in extract_daily_sheet(ws, ["UU", "PV"]):
                traffic_rows.append((dl_date, d, uu, pv))
            summaries[dl_date] = extract_summary_block(ws)

        if "デバイス別の使用状況" in wb.sheetnames:
            found_any_sheet = True
            ws = wb["デバイス別の使用状況"]
            if dl_date is None:
                dl_date = parse_download_date_or_default(ws, fallback_date)
            for row in extract_daily_sheet(ws, ["desktop", "mobileapp", "mobileweb", "tablet", "other"]):
                device_rows.append((dl_date, *row))

        if "人気のあるコンテンツ" in wb.sheetnames:
            found_any_sheet = True
            ws = wb["人気のあるコンテンツ"]
            if dl_date is None:
                dl_date = parse_download_date_or_default(ws, fallback_date)
            for c_name, typ, uu7, pv7 in extract_content_sheet(ws):
                content_rows.append((dl_date, c_name, typ, uu7, pv7))

        if "時間別の使用状況" in wb.sheetnames:
            found_any_sheet = True
            ws = wb["時間別の使用状況"]
            if dl_date is None:
                dl_date = parse_download_date_or_default(ws, fallback_date)
            rows = extract_hourly_sheet(ws)
            if latest_dl_date is None or (dl_date and dl_date >= latest_dl_date):
                latest_hourly = rows

        if not found_any_sheet:
            file_warnings.append(
                (name, "想定するシート（全体的なトラフィック 等）が見つかりませんでした。")
            )

        if dl_date is not None and (latest_dl_date is None or dl_date >= latest_dl_date):
            latest_dl_date = dl_date

    traffic_df = pd.DataFrame(traffic_rows, columns=["取得日", "日付", "UU", "PV"])
    device_df = pd.DataFrame(device_rows, columns=["取得日", "日付", "desktop", "mobileapp", "mobileweb", "tablet", "other"])
    content_df = pd.DataFrame(content_rows, columns=["取得日", "コンテンツ", "種類", "UU7", "PV7"])
    hourly_df = pd.DataFrame(latest_hourly or [], columns=["時間帯", "平均UU_7日"])

    def dedup_by_date(df):
        if df.empty:
            return df
        df = df.sort_values(["日付", "取得日"])
        return df.drop_duplicates(subset="日付", keep="last").sort_values("日付").reset_index(drop=True)

    traffic_df = dedup_by_date(traffic_df)
    device_df = dedup_by_date(device_df)

    latest_summary = summaries.get(latest_dl_date, {})

    gaps = []
    if not traffic_df.empty:
        dates_sorted = sorted(traffic_df["日付"].unique())
        for prev, cur in zip(dates_sorted, dates_sorted[1:]):
            gap_days = (cur - prev).days
            if gap_days > 1:
                gaps.append((prev, cur, gap_days))

    return {
        "traffic": traffic_df,
        "device": device_df,
        "content": content_df,
        "hourly": hourly_df,
        "latest_dl_date": latest_dl_date,
        "latest_summary": latest_summary,
        "n_files": len(sources),
        "gaps": gaps,
        "file_warnings": file_warnings,
    }


# ----------------------------------------------------------------------------
# グラフ生成（PNGをbase64化してHTMLに埋め込む＝外部ファイル・CDN不要）
# ----------------------------------------------------------------------------
def fig_to_base64(fig):
    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight")
    plt.close(fig)
    return base64.b64encode(buf.getvalue()).decode("ascii")


def chart_traffic_trend(traffic_df):
    n_days = max((traffic_df["日付"].max() - traffic_df["日付"].min()).days, 1)
    width = min(max(11, n_days / 12), 30)  # 期間が長いほど横に広げる（上限あり）

    fig, ax1 = plt.subplots(figsize=(width, 4))
    ax1.plot(traffic_df["日付"], traffic_df["PV"], color="#305496", linewidth=1,
              alpha=0.4, label="サイトの閲覧数(PV) 日次")
    if n_days >= 21:
        pv_ma = traffic_df["PV"].rolling(7, min_periods=1).mean()
        ax1.plot(traffic_df["日付"], pv_ma, color="#305496", linewidth=2,
                  label="サイトの閲覧数(PV) 7日移動平均")
    ax1.set_ylabel("サイトの閲覧数(PV)")

    ax2 = ax1.twinx()
    ax2.plot(traffic_df["日付"], traffic_df["UU"], color="#ED7D31", linewidth=1,
              alpha=0.35, label="重複しない閲覧者数(UU) 日次")
    if n_days >= 21:
        uu_ma = traffic_df["UU"].rolling(7, min_periods=1).mean()
        ax2.plot(traffic_df["日付"], uu_ma, color="#ED7D31", linewidth=2,
                  label="重複しない閲覧者数(UU) 7日移動平均")
    ax2.set_ylabel("重複しない閲覧者数(UU)")

    locator = mdates.AutoDateLocator(minticks=6, maxticks=14)
    ax1.xaxis.set_major_locator(locator)
    ax1.xaxis.set_major_formatter(mdates.ConciseDateFormatter(locator))

    fig.suptitle(f"日次トラフィック推移（{traffic_df['日付'].min()} 〜 {traffic_df['日付'].max()}、{len(traffic_df)}日分）")
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc="upper left", fontsize=8, ncol=2)
    fig.tight_layout()
    return fig_to_base64(fig)


def chart_device_trend(device_df):
    n_days = max((device_df["日付"].max() - device_df["日付"].min()).days, 1)
    width = min(max(11, n_days / 12), 30)

    fig, ax = plt.subplots(figsize=(width, 4))
    cols = ["desktop", "mobileapp", "mobileweb", "tablet", "other"]
    labels = ["デスクトップ", "モバイルアプリ", "モバイルウェブ", "タブレット", "その他"]
    ax.stackplot(device_df["日付"], *(device_df[c] for c in cols), labels=labels)

    locator = mdates.AutoDateLocator(minticks=6, maxticks=14)
    ax.xaxis.set_major_locator(locator)
    ax.xaxis.set_major_formatter(mdates.ConciseDateFormatter(locator))

    ax.set_title(f"デバイス別 閲覧数推移（積み上げ、{device_df['日付'].min()} 〜 {device_df['日付'].max()}）")
    ax.legend(loc="upper left", fontsize=8, ncol=5)
    fig.tight_layout()
    return fig_to_base64(fig)


def chart_cumulative_pv(traffic_df):
    """指定期間内の日次PVを積み上げた『累積PV』の推移を描画する。"""
    n_days = max((traffic_df["日付"].max() - traffic_df["日付"].min()).days, 1)
    width = min(max(11, n_days / 12), 30)

    cum_pv = traffic_df["PV"].cumsum()

    fig, ax = plt.subplots(figsize=(width, 4))
    ax.plot(traffic_df["日付"], cum_pv, color="#375623", linewidth=2)
    ax.fill_between(traffic_df["日付"], cum_pv, color="#375623", alpha=0.12)
    ax.set_ylabel("累積PV")

    locator = mdates.AutoDateLocator(minticks=6, maxticks=14)
    ax.xaxis.set_major_locator(locator)
    ax.xaxis.set_major_formatter(mdates.ConciseDateFormatter(locator))

    total_pv = int(cum_pv.iloc[-1]) if len(cum_pv) else 0
    ax.set_title(
        f"累積PV（{traffic_df['日付'].min()} 〜 {traffic_df['日付'].max()}、"
        f"期間内合計 {total_pv:,} PV）"
    )
    fig.tight_layout()
    return fig_to_base64(fig)


PERIODS = [
    ("1週間", 7),
    ("1ヶ月", 30),
    ("3ヶ月", 91),
    ("6ヶ月", 182),
    ("全期間", None),
]


def filter_period(df, date_col, days):
    """dfを『日付列の最大値からdays日前』以降に絞り込む。days=Noneなら全期間。"""
    if df.empty or days is None:
        return df
    end = df[date_col].max()
    start = end - datetime.timedelta(days=days - 1)
    return df[df[date_col] >= start]


def chart_top_content(content_df, latest_dl_date, top_n=10):
    latest = content_df[content_df["取得日"] == latest_dl_date].sort_values("PV7", ascending=False).head(top_n)
    fig, ax = plt.subplots(figsize=(9, 5))
    ax.barh(latest["コンテンツ"][::-1], latest["PV7"][::-1], color="#4472C4")
    ax.set_title(f"人気コンテンツ TOP{top_n}（直近7日・{latest_dl_date}時点）")
    ax.set_xlabel("閲覧数")
    return fig_to_base64(fig)


def chart_top_content_range(content_df, start_date, end_date, top_n=10):
    """指定期間内に含まれる週次スナップショット（取得日）を対象に、コンテンツごとの
    『過去7日間の閲覧数』を合算してランキングする（参考値：週次スナップショットの重なりにより
    厳密な期間合計ではないが、期間内での相対的な人気度の目安として利用できる）。"""
    sub = content_df[(content_df["取得日"] >= start_date) & (content_df["取得日"] <= end_date)]
    if sub.empty:
        return None
    agg = sub.groupby("コンテンツ", as_index=False)["PV7"].sum().sort_values("PV7", ascending=False).head(top_n)
    fig, ax = plt.subplots(figsize=(9, 5))
    ax.barh(agg["コンテンツ"][::-1], agg["PV7"][::-1], color="#4472C4")
    ax.set_title(f"人気コンテンツ TOP{top_n}（{start_date} 〜 {end_date}の週次スナップショット合計・参考値）")
    ax.set_xlabel("閲覧数（週次スナップショットの合算）")
    return fig_to_base64(fig)


def chart_hourly_heatmap(hourly_df):
    if hourly_df.empty:
        return None
    days_order = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    day_jp = {"Sunday": "日", "Monday": "月", "Tuesday": "火", "Wednesday": "水",
              "Thursday": "木", "Friday": "金", "Saturday": "土"}
    grid = pd.DataFrame(index=days_order, columns=range(24), dtype=float)
    for _, row in hourly_df.iterrows():
        m = re.match(r"(\w+) (\d+) (AM|PM) - ", row["時間帯"])
        if not m:
            continue
        day, hour, ampm = m.group(1), int(m.group(2)), m.group(3)
        h24 = (hour % 12) + (12 if ampm == "PM" else 0)
        if day in grid.index:
            grid.loc[day, h24] = row["平均UU_7日"]
    grid = grid.fillna(0)
    fig, ax = plt.subplots(figsize=(11, 3.5))
    im = ax.imshow(grid.values, aspect="auto", cmap="Blues")
    ax.set_yticks(range(len(days_order)))
    ax.set_yticklabels([day_jp[d] for d in days_order])
    ax.set_xticks(range(0, 24, 2))
    ax.set_xticklabels([f"{h}時" for h in range(0, 24, 2)])
    ax.set_title("曜日×時間帯 平均閲覧者数（直近7日平均）")
    fig.colorbar(im, ax=ax, shrink=0.7, label="平均UU")
    return fig_to_base64(fig)


# ----------------------------------------------------------------------------
# クライアント側（ブラウザ）で任意期間のグラフを再描画するための生データJSON化
# ----------------------------------------------------------------------------
def _num_or_none(v):
    """NaN/None を JSON の null に変換し、それ以外は float 化する。"""
    if v is None:
        return None
    try:
        if isinstance(v, float) and math.isnan(v):
            return None
    except TypeError:
        return None
    return float(v)


def traffic_records(df):
    """日次トラフィック（日付/PV/UU）をJSONシリアライズ可能なlist of dictへ変換。"""
    out = []
    for _, r in df.iterrows():
        out.append({
            "date": r["日付"].isoformat(),
            "pv": _num_or_none(r["PV"]),
            "uu": _num_or_none(r["UU"]),
        })
    return out


def device_records(df):
    """日次デバイス別使用状況をJSONシリアライズ可能なlist of dictへ変換。"""
    out = []
    for _, r in df.iterrows():
        out.append({
            "date": r["日付"].isoformat(),
            "desktop": _num_or_none(r["desktop"]),
            "mobileapp": _num_or_none(r["mobileapp"]),
            "mobileweb": _num_or_none(r["mobileweb"]),
            "tablet": _num_or_none(r["tablet"]),
            "other": _num_or_none(r["other"]),
        })
    return out


def content_records(df):
    """人気コンテンツの週次スナップショット（取得日ごとのPV7/UU7）をJSON化。
    カスタム期間選択時に、期間内の取得日スナップショットのみを合算してTOP10を作る。"""
    out = []
    for _, r in df.iterrows():
        out.append({
            "date": r["取得日"].isoformat() if hasattr(r["取得日"], "isoformat") else str(r["取得日"]),
            "name": r["コンテンツ"],
            "type": r["種類"],
            "uu7": _num_or_none(r["UU7"]),
            "pv7": _num_or_none(r["PV7"]),
        })
    return out


# ----------------------------------------------------------------------------
# カスタム期間グラフ描画用 JavaScript（ブラウザ側でSVGを再描画する）
# ----------------------------------------------------------------------------
CUSTOM_PERIOD_CHART_JS = """
<script>
/* ============================================================================
 * カスタム期間（ユーザー指定の開始日〜終了日）用グラフ描画エンジン
 * ----------------------------------------------------------------------------
 * サーバー側（Python/matplotlib）で事前生成する5つの固定期間パネルとは別に、
 * ブラウザ上で完結する軽量SVG描画で、任意の日付範囲のグラフをその場で作る。
 * 外部ライブラリ・CDNには依存しない（自己完結型HTMLの方針を維持）。
 * ==========================================================================*/
(function () {
  "use strict";

  var SVG_NS = "http://www.w3.org/2000/svg";
  var COLOR_PV = "#305496";
  var COLOR_UU = "#ED7D31";
  var COLOR_CUM = "#375623";
  var COLOR_CONTENT = "#4472C4";
  var DEVICE_COLORS = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd"];
  var DEVICE_COLS = ["desktop", "mobileapp", "mobileweb", "tablet", "other"];
  var DEVICE_LABELS = ["デスクトップ", "モバイルアプリ", "モバイルウェブ", "タブレット", "その他"];

  function loadRawData() {
    var tag = document.getElementById("dashboard-raw-data");
    if (!tag) return { traffic: [], device: [], content: [] };
    try {
      return JSON.parse(tag.textContent);
    } catch (e) {
      console.error("ダッシュボードの生データの読み込みに失敗しました:", e);
      return { traffic: [], device: [], content: [] };
    }
  }

  var RAW_DATA = loadRawData();

  function parseISO(s) {
    // "YYYY-MM-DD" をローカル日付として解釈（UTCずれ回避のため T00:00:00 を付与）
    return new Date(s + "T00:00:00");
  }

  function fmtDateShort(d, rangeDays) {
    var y = d.getFullYear(), m = d.getMonth() + 1, day = d.getDate();
    if (rangeDays > 400) {
      return y + "/" + m + "/" + day;
    }
    return m + "/" + day;
  }

  function fmtInt(n) {
    if (n === null || n === undefined || isNaN(n)) return "-";
    return Math.round(n).toLocaleString("ja-JP");
  }

  function fmtFloat1(n) {
    if (n === null || n === undefined || isNaN(n)) return "-";
    return (Math.round(n * 10) / 10).toLocaleString("ja-JP");
  }

  function svgEl(tag, attrs) {
    var e = document.createElementNS(SVG_NS, tag);
    if (attrs) {
      for (var k in attrs) {
        if (Object.prototype.hasOwnProperty.call(attrs, k)) {
          e.setAttribute(k, attrs[k]);
        }
      }
    }
    return e;
  }

  function scaleLinear(domain, range) {
    var d0 = domain[0], d1 = domain[1], r0 = range[0], r1 = range[1];
    var span = (d1 - d0);
    var m = span === 0 ? 0 : (r1 - r0) / span;
    return function (v) {
      return r0 + (v - d0) * m;
    };
  }

  function niceStep(rawStep) {
    if (rawStep <= 0) return 1;
    var pow10 = Math.pow(10, Math.floor(Math.log10(rawStep)));
    var f = rawStep / pow10;
    var nf;
    if (f <= 1) nf = 1;
    else if (f <= 2) nf = 2;
    else if (f <= 5) nf = 5;
    else nf = 10;
    return nf * pow10;
  }

  function yTicks(maxVal, count) {
    if (!isFinite(maxVal) || maxVal <= 0) maxVal = 1;
    var step = niceStep(maxVal / count);
    var top = Math.ceil(maxVal / step) * step;
    var ticks = [];
    for (var v = 0; v <= top + 1e-9; v += step) ticks.push(Math.round(v * 1000) / 1000);
    return ticks;
  }

  function clearAndMsg(container, msg) {
    container.innerHTML = "";
    var p = document.createElement("p");
    p.className = "note";
    p.textContent = msg;
    container.appendChild(p);
  }

  function rollingMean(values, window) {
    var out = new Array(values.length);
    var sum = 0;
    var q = [];
    for (var i = 0; i < values.length; i++) {
      q.push(values[i]);
      sum += values[i];
      if (q.length > window) sum -= q.shift();
      out[i] = sum / q.length;
    }
    return out;
  }

  /* ------------------------------------------------------------------------
   * 日次トラフィック推移（PV/UU 二軸折れ線）
   * ----------------------------------------------------------------------*/
  function renderTrendChart(container, rows) {
    container.innerHTML = "";
    if (rows.length === 0) {
      clearAndMsg(container, "データがありませんでした。");
      return;
    }
    var W = 960, H = 380, ML = 60, MR = 60, MT = 34, MB = 34;
    var innerW = W - ML - MR, innerH = H - MT - MB;

    var dates = rows.map(function (r) { return parseISO(r.date); });
    var pv = rows.map(function (r) { return r.pv || 0; });
    var uu = rows.map(function (r) { return r.uu || 0; });
    var t0 = dates[0].getTime(), t1 = dates[dates.length - 1].getTime();
    if (t0 === t1) { t0 -= 86400000; t1 += 86400000; }
    var rangeDays = Math.round((t1 - t0) / 86400000);

    var x = scaleLinear([t0, t1], [ML, ML + innerW]);
    var pvTicks = yTicks(Math.max.apply(null, pv), 5);
    var uuTicks = yTicks(Math.max.apply(null, uu), 5);
    var yPv = scaleLinear([0, pvTicks[pvTicks.length - 1]], [MT + innerH, MT]);
    var yUu = scaleLinear([0, uuTicks[uuTicks.length - 1]], [MT + innerH, MT]);

    var svg = svgEl("svg", { viewBox: "0 0 " + W + " " + H, width: "100%", height: "auto", style: "max-width:100%;" });

    // グリッド線・左Y軸（PV）
    pvTicks.forEach(function (v) {
      var yy = yPv(v);
      svg.appendChild(svgEl("line", { x1: ML, x2: ML + innerW, y1: yy, y2: yy, stroke: "#e6e6e6", "stroke-width": 1 }));
      var t = svgEl("text", { x: ML - 8, y: yy + 4, "text-anchor": "end", "font-size": 11, fill: COLOR_PV });
      t.textContent = fmtInt(v);
      svg.appendChild(t);
    });
    // 右Y軸（UU）
    uuTicks.forEach(function (v) {
      var yy = yUu(v);
      var t = svgEl("text", { x: ML + innerW + 8, y: yy + 4, "text-anchor": "start", "font-size": 11, fill: COLOR_UU });
      t.textContent = fmtInt(v);
      svg.appendChild(t);
    });

    // X軸目盛
    var nTicks = Math.min(8, dates.length);
    for (var i = 0; i < nTicks; i++) {
      var frac = nTicks === 1 ? 0 : i / (nTicks - 1);
      var tx = t0 + frac * (t1 - t0);
      var xx = x(tx);
      svg.appendChild(svgEl("line", { x1: xx, x2: xx, y1: MT, y2: MT + innerH, stroke: "#f0f0f0", "stroke-width": 1 }));
      var t = svgEl("text", { x: xx, y: MT + innerH + 18, "text-anchor": "middle", "font-size": 11, fill: "#555" });
      t.textContent = fmtDateShort(new Date(tx), rangeDays);
      svg.appendChild(t);
    }

    function pathFor(values, yScale) {
      var d = "";
      for (var i = 0; i < values.length; i++) {
        var xx = x(dates[i].getTime());
        var yy = yScale(values[i]);
        d += (i === 0 ? "M" : "L") + xx.toFixed(1) + "," + yy.toFixed(1) + " ";
      }
      return d;
    }

    // 日次（薄い線）
    svg.appendChild(svgEl("path", { d: pathFor(pv, yPv), fill: "none", stroke: COLOR_PV, "stroke-width": 1, "stroke-opacity": 0.4 }));
    svg.appendChild(svgEl("path", { d: pathFor(uu, yUu), fill: "none", stroke: COLOR_UU, "stroke-width": 1, "stroke-opacity": 0.35 }));

    var showMA = rangeDays >= 21;
    if (showMA) {
      var pvMa = rollingMean(pv, 7);
      var uuMa = rollingMean(uu, 7);
      svg.appendChild(svgEl("path", { d: pathFor(pvMa, yPv), fill: "none", stroke: COLOR_PV, "stroke-width": 2 }));
      svg.appendChild(svgEl("path", { d: pathFor(uuMa, yUu), fill: "none", stroke: COLOR_UU, "stroke-width": 2 }));
    }

    // 軸ラベル
    var lblPv = svgEl("text", { x: 14, y: MT - 12, "font-size": 12, fill: COLOR_PV });
    lblPv.textContent = "サイトの閲覧数(PV)";
    svg.appendChild(lblPv);
    var lblUu = svgEl("text", { x: W - 14, y: MT - 12, "text-anchor": "end", "font-size": 12, fill: COLOR_UU });
    lblUu.textContent = "重複しない閲覧者数(UU)";
    svg.appendChild(lblUu);

    // タイトル
    var title = svgEl("text", { x: W / 2, y: 16, "text-anchor": "middle", "font-size": 13, fill: "#333" });
    title.textContent = "日次トラフィック推移（" + rows[0].date + " 〜 " + rows[rows.length - 1].date + "、" + rows.length + "日分）";
    svg.appendChild(title);

    // 凡例
    var legendItems = showMA
      ? [["PV 日次", COLOR_PV, 0.4], ["PV 7日移動平均", COLOR_PV, 1], ["UU 日次", COLOR_UU, 0.35], ["UU 7日移動平均", COLOR_UU, 1]]
      : [["PV 日次", COLOR_PV, 0.4], ["UU 日次", COLOR_UU, 0.35]];
    var lx = ML + 6, ly = MT + 14;
    legendItems.forEach(function (it, idx) {
      var ypos = ly + idx * 15;
      svg.appendChild(svgEl("line", { x1: lx, x2: lx + 18, y1: ypos, y2: ypos, stroke: it[1], "stroke-width": 2, "stroke-opacity": it[2] }));
      var t = svgEl("text", { x: lx + 24, y: ypos + 4, "font-size": 10, fill: "#333" });
      t.textContent = it[0];
      svg.appendChild(t);
    });

    container.appendChild(svg);
  }

  /* ------------------------------------------------------------------------
   * 累積PV
   * ----------------------------------------------------------------------*/
  function renderCumulativeChart(container, rows) {
    container.innerHTML = "";
    if (rows.length === 0) {
      clearAndMsg(container, "データがありませんでした。");
      return;
    }
    var W = 960, H = 380, ML = 70, MR = 30, MT = 34, MB = 34;
    var innerW = W - ML - MR, innerH = H - MT - MB;

    var dates = rows.map(function (r) { return parseISO(r.date); });
    var pv = rows.map(function (r) { return r.pv || 0; });
    var cum = [];
    var acc = 0;
    for (var i = 0; i < pv.length; i++) { acc += pv[i]; cum.push(acc); }

    var t0 = dates[0].getTime(), t1 = dates[dates.length - 1].getTime();
    if (t0 === t1) { t0 -= 86400000; t1 += 86400000; }
    var rangeDays = Math.round((t1 - t0) / 86400000);

    var x = scaleLinear([t0, t1], [ML, ML + innerW]);
    var ticks = yTicks(cum[cum.length - 1], 5);
    var y = scaleLinear([0, ticks[ticks.length - 1]], [MT + innerH, MT]);

    var svg = svgEl("svg", { viewBox: "0 0 " + W + " " + H, width: "100%", height: "auto", style: "max-width:100%;" });

    ticks.forEach(function (v) {
      var yy = y(v);
      svg.appendChild(svgEl("line", { x1: ML, x2: ML + innerW, y1: yy, y2: yy, stroke: "#e6e6e6", "stroke-width": 1 }));
      var t = svgEl("text", { x: ML - 8, y: yy + 4, "text-anchor": "end", "font-size": 11, fill: "#333" });
      t.textContent = fmtInt(v);
      svg.appendChild(t);
    });

    var nTicks = Math.min(8, dates.length);
    for (var i2 = 0; i2 < nTicks; i2++) {
      var frac = nTicks === 1 ? 0 : i2 / (nTicks - 1);
      var tx = t0 + frac * (t1 - t0);
      var xx = x(tx);
      var t3 = svgEl("text", { x: xx, y: MT + innerH + 18, "text-anchor": "middle", "font-size": 11, fill: "#555" });
      t3.textContent = fmtDateShort(new Date(tx), rangeDays);
      svg.appendChild(t3);
    }

    var linePath = "", areaPath = "";
    for (var i3 = 0; i3 < cum.length; i3++) {
      var xx3 = x(dates[i3].getTime());
      var yy3 = y(cum[i3]);
      linePath += (i3 === 0 ? "M" : "L") + xx3.toFixed(1) + "," + yy3.toFixed(1) + " ";
    }
    areaPath = "M" + x(dates[0].getTime()).toFixed(1) + "," + (MT + innerH) + " ";
    for (var i4 = 0; i4 < cum.length; i4++) {
      areaPath += "L" + x(dates[i4].getTime()).toFixed(1) + "," + y(cum[i4]).toFixed(1) + " ";
    }
    areaPath += "L" + x(dates[cum.length - 1].getTime()).toFixed(1) + "," + (MT + innerH) + " Z";

    svg.appendChild(svgEl("path", { d: areaPath, fill: COLOR_CUM, "fill-opacity": 0.12, stroke: "none" }));
    svg.appendChild(svgEl("path", { d: linePath, fill: "none", stroke: COLOR_CUM, "stroke-width": 2 }));

    var title = svgEl("text", { x: W / 2, y: 16, "text-anchor": "middle", "font-size": 13, fill: "#333" });
    title.textContent = "累積PV（" + rows[0].date + " 〜 " + rows[rows.length - 1].date + "、期間内合計 " + fmtInt(cum[cum.length - 1]) + " PV）";
    svg.appendChild(title);

    var ylabel = svgEl("text", { x: 14, y: MT - 12, "font-size": 12, fill: "#333" });
    ylabel.textContent = "累積PV";
    svg.appendChild(ylabel);

    container.appendChild(svg);
  }

  /* ------------------------------------------------------------------------
   * デバイス別 閲覧数推移（積み上げ面グラフ）
   * ----------------------------------------------------------------------*/
  function renderDeviceChart(container, rows) {
    container.innerHTML = "";
    if (rows.length === 0) {
      clearAndMsg(container, "データがありませんでした。");
      return;
    }
    var W = 960, H = 380, ML = 60, MR = 30, MT = 46, MB = 34;
    var innerW = W - ML - MR, innerH = H - MT - MB;

    var dates = rows.map(function (r) { return parseISO(r.date); });
    var series = DEVICE_COLS.map(function (c) { return rows.map(function (r) { return r[c] || 0; }); });
    var totals = rows.map(function (_, i) {
      var s = 0;
      for (var k = 0; k < series.length; k++) s += series[k][i];
      return s;
    });

    var t0 = dates[0].getTime(), t1 = dates[dates.length - 1].getTime();
    if (t0 === t1) { t0 -= 86400000; t1 += 86400000; }
    var rangeDays = Math.round((t1 - t0) / 86400000);

    var x = scaleLinear([t0, t1], [ML, ML + innerW]);
    var ticks = yTicks(Math.max.apply(null, totals), 5);
    var y = scaleLinear([0, ticks[ticks.length - 1]], [MT + innerH, MT]);

    var svg = svgEl("svg", { viewBox: "0 0 " + W + " " + H, width: "100%", height: "auto", style: "max-width:100%;" });

    ticks.forEach(function (v) {
      var yy = y(v);
      svg.appendChild(svgEl("line", { x1: ML, x2: ML + innerW, y1: yy, y2: yy, stroke: "#e6e6e6", "stroke-width": 1 }));
      var t = svgEl("text", { x: ML - 8, y: yy + 4, "text-anchor": "end", "font-size": 11, fill: "#333" });
      t.textContent = fmtInt(v);
      svg.appendChild(t);
    });

    var nTicks = Math.min(8, dates.length);
    for (var i = 0; i < nTicks; i++) {
      var frac = nTicks === 1 ? 0 : i / (nTicks - 1);
      var tx = t0 + frac * (t1 - t0);
      var xx = x(tx);
      var t2 = svgEl("text", { x: xx, y: MT + innerH + 18, "text-anchor": "middle", "font-size": 11, fill: "#555" });
      t2.textContent = fmtDateShort(new Date(tx), rangeDays);
      svg.appendChild(t2);
    }

    // 積み上げ用の累積配列を作る
    var stackBase = rows.map(function () { return 0; });
    var cumSeries = [];
    for (var s = 0; s < series.length; s++) {
      var top = [];
      for (var i2 = 0; i2 < rows.length; i2++) {
        stackBase[i2] += series[s][i2];
        top.push(stackBase[i2]);
      }
      cumSeries.push(top.slice());
    }
    var prevSeries = cumSeries.map(function (_, idx) {
      return idx === 0 ? rows.map(function () { return 0; }) : cumSeries[idx - 1];
    });

    for (var s2 = 0; s2 < series.length; s2++) {
      var top2 = cumSeries[s2], bot2 = prevSeries[s2];
      var d = "M" + x(dates[0].getTime()).toFixed(1) + "," + y(bot2[0]).toFixed(1) + " ";
      for (var i3 = 0; i3 < rows.length; i3++) d += "L" + x(dates[i3].getTime()).toFixed(1) + "," + y(top2[i3]).toFixed(1) + " ";
      for (var i4 = rows.length - 1; i4 >= 0; i4--) d += "L" + x(dates[i4].getTime()).toFixed(1) + "," + y(bot2[i4]).toFixed(1) + " ";
      d += "Z";
      svg.appendChild(svgEl("path", { d: d, fill: DEVICE_COLORS[s2], "fill-opacity": 0.85, stroke: "none" }));
    }

    var title = svgEl("text", { x: W / 2, y: 16, "text-anchor": "middle", "font-size": 13, fill: "#333" });
    title.textContent = "デバイス別 閲覧数推移（積み上げ、" + rows[0].date + " 〜 " + rows[rows.length - 1].date + "）";
    svg.appendChild(title);

    // 凡例（横並び）
    var lx = ML, ly = MT - 22;
    DEVICE_LABELS.forEach(function (label, idx) {
      var gx = lx + idx * 150;
      svg.appendChild(svgEl("rect", { x: gx, y: ly - 9, width: 12, height: 12, fill: DEVICE_COLORS[idx] }));
      var t = svgEl("text", { x: gx + 16, y: ly + 1, "font-size": 10, fill: "#333" });
      t.textContent = label;
      svg.appendChild(t);
    });

    container.appendChild(svg);
  }

  /* ------------------------------------------------------------------------
   * 人気コンテンツ TOP10（期間内の週次スナップショットを合算・参考値）
   * ----------------------------------------------------------------------*/
  function renderContentChart(container, contentRows, startDate, endDate, topN) {
    container.innerHTML = "";
    topN = topN || 10;
    if (contentRows.length === 0) {
      clearAndMsg(container, "データがありませんでした。");
      return;
    }
    var agg = {};
    contentRows.forEach(function (r) {
      var name = r.name || "(不明)";
      agg[name] = (agg[name] || 0) + (r.pv7 || 0);
    });
    var entries = Object.keys(agg).map(function (k) { return [k, agg[k]]; });
    entries.sort(function (a, b) { return b[1] - a[1]; });
    entries = entries.slice(0, topN).reverse(); // 下から積み上げ表示のため逆順

    var W = 900, rowH = 30, MT = 40, MB = 40, ML = 260, MR = 40;
    var H = MT + MB + entries.length * rowH;
    var innerW = W - ML - MR, innerH = entries.length * rowH;

    var maxVal = Math.max.apply(null, entries.map(function (e) { return e[1]; }));
    var ticks = yTicks(maxVal, 5);
    var x = scaleLinear([0, ticks[ticks.length - 1]], [ML, ML + innerW]);

    var svg = svgEl("svg", { viewBox: "0 0 " + W + " " + H, width: "100%", height: "auto", style: "max-width:100%;" });

    ticks.forEach(function (v) {
      var xx = x(v);
      svg.appendChild(svgEl("line", { x1: xx, x2: xx, y1: MT, y2: MT + innerH, stroke: "#e6e6e6", "stroke-width": 1 }));
      var t = svgEl("text", { x: xx, y: MT + innerH + 16, "text-anchor": "middle", "font-size": 10, fill: "#555" });
      t.textContent = fmtInt(v);
      svg.appendChild(t);
    });

    entries.forEach(function (e, idx) {
      var yTop = MT + idx * rowH + rowH * 0.15;
      var barH = rowH * 0.7;
      var w = x(e[1]) - ML;
      svg.appendChild(svgEl("rect", { x: ML, y: yTop, width: Math.max(w, 0), height: barH, fill: COLOR_CONTENT }));
      var name = e[0].length > 34 ? e[0].slice(0, 33) + "…" : e[0];
      var lbl = svgEl("text", { x: ML - 8, y: yTop + barH / 2 + 4, "text-anchor": "end", "font-size": 11, fill: "#333" });
      lbl.textContent = name;
      svg.appendChild(lbl);
      var val = svgEl("text", { x: x(e[1]) + 6, y: yTop + barH / 2 + 4, "text-anchor": "start", "font-size": 10, fill: "#333" });
      val.textContent = fmtInt(e[1]);
      svg.appendChild(val);
    });

    var title = svgEl("text", { x: W / 2, y: 18, "text-anchor": "middle", "font-size": 13, fill: "#333" });
    title.textContent = "人気コンテンツ TOP" + topN + "（" + startDate + " 〜 " + endDate + " の週次スナップショット合計・参考値）";
    svg.appendChild(title);

    var xlabel = svgEl("text", { x: ML + innerW / 2, y: H - 8, "text-anchor": "middle", "font-size": 11, fill: "#555" });
    xlabel.textContent = "閲覧数（週次スナップショットの合算）";
    svg.appendChild(xlabel);

    container.appendChild(svg);
  }

  /* ------------------------------------------------------------------------
   * カスタム期間の適用（開始日・終了日を読み取り、全グラフ＆サマリーを更新）
   * ----------------------------------------------------------------------*/
  function applyCustomPeriod() {
    var msgEl = document.getElementById("custom-range-msg");
    var startInput = document.getElementById("custom-start");
    var endInput = document.getElementById("custom-end");
    if (!startInput || !endInput) return;

    var startStr = startInput.value;
    var endStr = endInput.value;
    if (!startStr || !endStr) {
      msgEl.textContent = "開始日・終了日を指定してください。";
      return;
    }
    var startD = parseISO(startStr);
    var endD = parseISO(endStr);
    if (startD > endD) {
      msgEl.textContent = "開始日は終了日より前にしてください。";
      return;
    }
    msgEl.textContent = "";

    var trafficAll = RAW_DATA.traffic || [];
    var deviceAll = RAW_DATA.device || [];
    var contentAll = RAW_DATA.content || [];

    var trafficRows = trafficAll.filter(function (r) {
      var d = parseISO(r.date);
      return d >= startD && d <= endD;
    });
    var deviceRows = deviceAll.filter(function (r) {
      var d = parseISO(r.date);
      return d >= startD && d <= endD;
    });
    var contentRows = contentAll.filter(function (r) {
      var d = parseISO(r.date);
      return d >= startD && d <= endD;
    });

    // --- サマリー ---
    var summaryDiv = document.getElementById("custom-summary");
    if (summaryDiv) {
      if (trafficRows.length === 0) {
        summaryDiv.innerHTML = "<p class='note'>指定期間にデータがありませんでした。</p>";
      } else {
        var totalPv = trafficRows.reduce(function (s, r) { return s + (r.pv || 0); }, 0);
        var avgUu = trafficRows.reduce(function (s, r) { return s + (r.uu || 0); }, 0) / trafficRows.length;
        summaryDiv.innerHTML =
          "<table class='custom-summary-tbl'>" +
          "<tr><th>期間</th><th>対象日数</th><th>PV合計</th><th>UU 日次平均</th></tr>" +
          "<tr><td>" + startStr + " 〜 " + endStr + "</td><td>" + trafficRows.length + " 日</td>" +
          "<td>" + fmtInt(totalPv) + "</td><td>" + fmtFloat1(avgUu) + "</td></tr>" +
          "</table>" +
          "<p class='note'>※ PV合計は日次値の単純合計（正確な値）です。UUは日をまたいで重複しうるため、" +
          "参考値として「日次平均」を表示しています（正確な期間合計UUはSharePoint公式集計値をご確認ください）。</p>";
      }
    }

    renderTrendChart(document.getElementById("custom-chart-trend"), trafficRows);
    renderCumulativeChart(document.getElementById("custom-chart-cumulative"), trafficRows);
    renderDeviceChart(document.getElementById("custom-chart-device"), deviceRows);
    renderContentChart(document.getElementById("custom-chart-content"), contentRows, startStr, endStr, 10);
  }

  // グローバルに公開（onclick 属性から呼び出すため）
  window.applyCustomPeriod = applyCustomPeriod;
})();

</script>
"""


# ----------------------------------------------------------------------------
# HTML組み立て
# ----------------------------------------------------------------------------
def build_html(data, out_path=None):
    traffic_df, device_df, content_df = data["traffic"], data["device"], data["content"]
    latest_dl_date = data["latest_dl_date"]
    summary = data["latest_summary"]

    # --- 期間ごと（1ヶ月/3ヶ月/6ヶ月/全期間）にグラフを作り分ける ---
    period_panels = []  # (label, html_id, inner_html)
    for label, days in PERIODS:
        t_sub = filter_period(traffic_df, "日付", days)
        d_sub = filter_period(device_df, "日付", days)

        img_trend = chart_traffic_trend(t_sub) if not t_sub.empty else None
        img_cum_pv = chart_cumulative_pv(t_sub) if not t_sub.empty else None
        img_device = chart_device_trend(d_sub) if not d_sub.empty else None

        if not t_sub.empty:
            start_date, end_date = t_sub["日付"].min(), t_sub["日付"].max()
            img_content = chart_top_content_range(content_df, start_date, end_date) if not content_df.empty else None
        else:
            img_content = None

        def img_block(title, b64):
            if b64 is None:
                return f"<h2>{title}</h2><p class='note'>データがありませんでした。</p>"
            return f"<h2>{title}</h2><img src='data:image/png;base64,{b64}' style='max-width:100%;'>"

        inner = (
            img_block("日次トラフィック推移", img_trend)
            + img_block("累積PV", img_cum_pv)
            + img_block("デバイス別 閲覧数推移", img_device)
            + img_block("人気コンテンツ TOP10", img_content)
        )
        html_id = f"period-{days if days else 'all'}"
        period_panels.append((label, html_id, inner))

    # 曜日×時間帯ヒートマップは SharePoint 側が常に「直近7日平均」しか提供しないため、
    # 期間切替の対象外（共通表示）とする。
    img_hourly = chart_hourly_heatmap(data["hourly"])

    # --- カスタム期間（ユーザーが開始日・終了日を自由に指定）用の生データをJSON化 ---
    # 事前生成画像とは異なり、ブラウザ側のJavaScriptで都度グラフを再描画するため、
    # 日次データをそのままクライアントに渡す。
    raw_data_json = json.dumps({
        "traffic": traffic_records(traffic_df),
        "device": device_records(device_df),
        "content": content_records(content_df),
    }, ensure_ascii=False)

    if not traffic_df.empty:
        data_min_date = traffic_df["日付"].min().isoformat()
        data_max_date = traffic_df["日付"].max().isoformat()
    else:
        data_min_date = data_max_date = ""

    def summary_row(label):
        uu, pv = summary.get(label, ("-", "-"))
        return f"<tr><td>{label}</td><td>{uu}</td><td>{pv}</td></tr>"

    summary_table = f"""
    <table class="tbl">
      <tr><th>期間</th><th>重複しない閲覧者数(UU)</th><th>サイトの閲覧数(PV)</th></tr>
      {summary_row('過去 7 日')}
      {summary_row('過去 30 日')}
      {summary_row('過去 90 日間')}
      {summary_row('すべての時間')}
    </table>
    <p class="note">※ 上表は最新ファイル（{latest_dl_date}）に記載された SharePoint 公式の集計値をそのまま表示しています
    （日次値の合算ではないため、UUも正確です）。下のグラフの期間切替とは連動していません。</p>
    """

    gaps = data.get("gaps", [])
    if gaps:
        items = "".join(f"<li>{p} 〜 {c}（{g-1}日分欠測）</li>" for p, c, g in gaps[:20])
        gap_warning_html = (
            f"<div style='background:#fce4d6;border:1px solid #e07b39;padding:10px;border-radius:4px;'>"
            f"<b>⚠ データに欠測期間が {len(gaps)} 件あります</b>（90日を超えてファイルを取得しなかった期間の可能性）。"
            f"<ul>{items}</ul></div>"
        )
    else:
        gap_warning_html = ""

    def hourly_block(b64):
        if b64 is None:
            return "<h2>曜日×時間帯 利用状況</h2><p class='note'>データがありませんでした。</p>"
        return (
            "<h2>曜日×時間帯 利用状況</h2>"
            "<p class='note'>※ SharePointの仕様上、常に「直近7日平均」のみが提供されるため、期間切替の対象外です。</p>"
            f"<img src='data:image/png;base64,{b64}' style='max-width:100%;'>"
        )

    # --- 期間切替ボタン & パネルのHTML/JS ---
    default_id = period_panels[-1][1]  # 既定表示は「全期間」
    buttons_html = "".join(
        f"<button class='period-btn{' active' if hid == default_id else ''}' "
        f"data-target='{hid}' onclick=\"showPeriod('{hid}')\">{label}</button>"
        for label, hid, _ in period_panels
    )
    # 「カスタム期間」ボタンを末尾に追加
    buttons_html += (
        "<button class='period-btn' data-target='period-custom' "
        "onclick=\"showPeriod('period-custom')\">カスタム期間</button>"
    )

    panels_html = "".join(
        f"<div class='period-panel' id='{hid}' style='display:{'block' if hid == default_id else 'none'};'>{inner}</div>"
        for _, hid, inner in period_panels
    )
    # 「カスタム期間」パネル（開始日・終了日をユーザーが指定し、JavaScriptでグラフを再描画）
    panels_html += f"""
    <div class="period-panel" id="period-custom" style="display:none;">
      <div class="custom-range-bar">
        <label>開始日: <input type="date" id="custom-start" min="{data_min_date}" max="{data_max_date}" value="{data_min_date}"></label>
        <label>終了日: <input type="date" id="custom-end" min="{data_min_date}" max="{data_max_date}" value="{data_max_date}"></label>
        <button class="apply-btn" onclick="applyCustomPeriod()">適用</button>
        <span id="custom-range-msg" class="note"></span>
      </div>
      <p class="note">対象データ範囲: {data_min_date} 〜 {data_max_date}（この範囲内で自由に期間を指定できます）</p>
      <div id="custom-summary"></div>
      <h2>日次トラフィック推移</h2>
      <div id="custom-chart-trend"></div>
      <h2>累積PV</h2>
      <div id="custom-chart-cumulative"></div>
      <h2>デバイス別 閲覧数推移</h2>
      <div id="custom-chart-device"></div>
      <h2>人気コンテンツ TOP10</h2>
      <div id="custom-chart-content"></div>
    </div>
    """

    chart_engine_js = CUSTOM_PERIOD_CHART_JS

    html = f"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="utf-8">
<title>サイト使用状況ダッシュボード</title>
<style>
  body {{ font-family: 'Meiryo', 'Yu Gothic', sans-serif; margin: 30px; color:#222; background:#f7f8fa; }}
  h1 {{ color:#305496; }}
  h2 {{ color:#305496; border-bottom: 2px solid #d9e2f3; padding-bottom:4px; margin-top:40px; }}
  .tbl {{ border-collapse: collapse; margin: 10px 0; }}
  .tbl th, .tbl td {{ border: 1px solid #ccc; padding: 6px 14px; text-align: right; }}
  .tbl th:first-child, .tbl td:first-child {{ text-align: left; }}
  .tbl th {{ background: #d9e2f3; }}
  .note {{ color:#777; font-size: 0.85em; }}
  .meta {{ color:#555; font-size:0.9em; }}
  .period-bar {{ margin-top: 30px; }}
  .period-btn {{
      font-family: inherit; font-size: 0.95em; padding: 8px 20px; margin-right: 8px;
      border: 1px solid #305496; background: #fff; color: #305496; border-radius: 20px;
      cursor: pointer; transition: 0.15s;
  }}
  .period-btn:hover {{ background: #d9e2f3; }}
  .period-btn.active {{ background: #305496; color: #fff; }}
  .custom-range-bar {{ margin: 14px 0; display:flex; align-items:center; flex-wrap:wrap; gap:16px; }}
  .custom-range-bar label {{ font-size:0.95em; color:#333; }}
  .custom-range-bar input[type=date] {{
      font-family: inherit; font-size:0.95em; padding:4px 8px; margin-left:6px;
      border:1px solid #bbb; border-radius:4px;
  }}
  .apply-btn {{
      font-family: inherit; font-size:0.95em; padding:7px 22px;
      border:none; background:#305496; color:#fff; border-radius:20px; cursor:pointer;
  }}
  .apply-btn:hover {{ background:#20385f; }}
  #custom-range-msg {{ color:#c0392b; }}
  .custom-summary-tbl {{ border-collapse: collapse; margin: 10px 0; }}
  .custom-summary-tbl th, .custom-summary-tbl td {{ border: 1px solid #ccc; padding: 6px 14px; text-align: right; }}
  .custom-summary-tbl th:first-child, .custom-summary-tbl td:first-child {{ text-align: left; }}
  .custom-summary-tbl th {{ background: #d9e2f3; }}
  .chart-legend {{ font-size:0.8em; }}
</style>
</head>
<body>
  <h1>サイト使用状況ダッシュボード</h1>
  <p class="meta">生成日時: {datetime.datetime.now():%Y-%m-%d %H:%M} ／ 読み込みファイル数: {data['n_files']} ／
     最新データ取得日: {latest_dl_date}</p>

  {gap_warning_html}

  <h2>サマリー</h2>
  {summary_table}

  <div class="period-bar">
    <span class="meta">表示期間：</span>
    {buttons_html}
  </div>
  {panels_html}

  {hourly_block(img_hourly)}

<script id="dashboard-raw-data" type="application/json">{raw_data_json}</script>

<script>
function showPeriod(id) {{
    document.querySelectorAll('.period-panel').forEach(function(el) {{
        el.style.display = (el.id === id) ? 'block' : 'none';
    }});
    document.querySelectorAll('.period-btn').forEach(function(btn) {{
        btn.classList.toggle('active', btn.getAttribute('data-target') === id);
    }});
    // 「カスタム期間」タブを初めて開いたときに、既定範囲（全期間）で一度描画しておく
    if (id === 'period-custom' && typeof window.__customRendered === 'undefined') {{
        window.__customRendered = true;
        applyCustomPeriod();
    }}
}}
</script>
{chart_engine_js}
</body>
</html>
"""
    if out_path is not None:
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(html)
        print(f"ダッシュボードを出力しました: {out_path}")
    return html


def build_html_string(data):
    """Streamlit（アップロード）版向け: ファイルに書き出さず、HTML文字列をそのまま返す。
    dashboard_maker.py はこの戻り値を st.download_button の data として使う。"""
    return build_html(data, out_path=None)


def main():
    if len(sys.argv) < 2:
        print("使い方: python build_dashboard.py <ログフォルダ> [出力HTML]")
        sys.exit(1)
    folder = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else "dashboard.html"

    data = load_all(folder)
    build_html(data, out_path)


if __name__ == "__main__":
    main()
